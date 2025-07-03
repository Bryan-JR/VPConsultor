from flask import Blueprint, request, jsonify, render_template, send_file
from src.models.Inventario import Inventario
from src.models.Referencias import Referencia
from src.models.UltimaFechaCompra import UltimaFechaCompra
from sqlalchemy import func, text
from src.db import Session, engine
import pandas as pd
import numpy as np
import os
import io
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, numbers
from openpyxl.workbook.defined_name import DefinedName
from datetime import datetime, time
from apscheduler.schedulers.background import BackgroundScheduler # type: ignore
controllerInventario = Blueprint('controllerInventario', __name__)
import socket
from functools import lru_cache
import locale
locale.setlocale(locale.LC_TIME, 'Spanish_Spain.1252')

ruta = r'\\10.0.0.96\Compartida\Planos'
inv = os.path.join(ruta, "Inventario.txt")


def datos(archivo):
    encodings = ['utf-8', 'ISO-8859-1', 'latin1']
    for encoding in encodings:
        try:
            return pd.read_csv(archivo, sep=';', encoding=encoding, on_bad_lines='skip', dtype={'REFERENCIA': str})
        except UnicodeDecodeError:
            continue
    raise ValueError("No se puede decodificar el archivo con codificaciones disponibles")


from sqlalchemy.orm import sessionmaker
from sqlalchemy import func

def actualizarInv():
    try:
        hora_inicio = time(7, 40)
        hora_fin = time(20, 40)
        hora_actual = datetime.now().time()
        init = False
        if(hora_inicio <= hora_actual <= hora_fin):
                inv = os.path.join(ruta, "Inventario.txt")
                #tam = os.path.getsize(r'\\10.0.0.100\serveriltda\Iltda\Soporte Compartida\Inventario.txt')
                df = datos(inv)
                NaN_ref = df['REFERENCIA'].isna().any()
                NaN_lin = df['LINEA'].isna().any()
                NaN_bd = df['BODEGA'].isna().any()
                NaN_exi = df['EXISTENCIA'].isna().any()
                NaN_cos = df['COSTO'].isna().any()
                
                
                if not NaN_ref and not NaN_bd and not NaN_lin and not NaN_exi and not NaN_cos:
                    conditions = [
                    (df['TIPO IVA'].between(41, 46)) | (df['TIPO IVA'].between(48, 54)),
                    (df['TIPO IVA'].isin([39, 47, 55, 56])),
                    (df['TIPO IVA'].isin([98, 99])),
                    ]
                    values = [1.19, 1.05, 1]
                    df['COSTO'] = (df['COSTO'] * np.select(conditions, values, default=1)).round(0).astype(int)
                    session = Session()
                    init = True
                    # Traer todos los registros existentes de la base de datos
                    inventarios_existentes = session.query(Inventario).all()
                    inventario_dict = {(inv.REFERENCIA, inv.BODEGA): inv for inv in inventarios_existentes}

                    #referencias_actualizadas = set((str(int(row['REFERENCIA'])), int(row['BODEGA'])) for _, row in df.iterrows())

                    # Iterar sobre las filas del DataFrame
                    for _, row in df.iterrows():
                        clave = (str(int(row['REFERENCIA'])), int(row['BODEGA']))
                        if clave in inventario_dict:
                            # Si el registro existe, actualizar EXISTENCIA y ESTADO
                            inventario_existente = inventario_dict[clave]
                            inventario_existente.LINEA = int(row['LINEA'])
                            inventario_existente.EXISTENCIA = int(row['EXISTENCIA'])
                            inventario_existente.PROVEEDOR = int(row['PROVEEDOR']) if pd.notna(row['PROVEEDOR']) else None
                            inventario_existente.COSTO = int(row['COSTO']) if pd.notna(row['COSTO']) else None
                        else:
                            # Si no existe, crear un nuevo registro (nueva referencia)
                            nuevo_inventario = Inventario(
                                REFERENCIA=str(int(row['REFERENCIA'])),
                                BODEGA=int(row['BODEGA']),
                                LINEA=int(row['LINEA']),
                                PROVEEDOR=int(row['PROVEEDOR']) if pd.notna(row['PROVEEDOR']) else None,
                                SUBGRUPO=int(row['SUBGRUPO']) if pd.notna(row['SUBGRUPO']) else None,
                                EXISTENCIA=int(row['EXISTENCIA']) ,
                                COSTO=int(row['COSTO']) if pd.notna(row['COSTO']) else None
                            )
                            session.add(nuevo_inventario)
                    
                    # referencias_existentes = set(inventario_dict.keys())
                    # referencias_a_eliminar = referencias_existentes - referencias_actualizadas
                    # print(referencias_a_eliminar)
                    
                    
                    session.commit()
                    total_existencia = session.query(func.sum(Inventario.EXISTENCIA)).scalar() or 0
                    msg("SERVER", f"Inventario actualizado correctamente. Total existencia: ---* {total_existencia} *---")
                    
    except Exception as e:
        print("Error en actualizar inventario: ", e)
        session.rollback()  # Revertir en caso de error
    finally:
        if init:
            session.close()

# METODO PARA ACTUALIZAR ULTIMA FECHA DE COMPRA
def actualizarFechaCompra():
    try:
        fechas = os.path.join(ruta, "MovCompras.txt")
        #tam = os.path.getsize(r'\\10.0.0.100\serveriltda\Iltda\Soporte Compartida\Inventario.txt')
        init = False
        df = datos(fechas)
        #df = df[(df['BODEGA'] == 2) | df['BODEGA'] == 1]
        NaN_ref = df['REFERENCIA'].isna().any()
        NaN_fec = df['FECHA'].isna().any()
        if not NaN_ref and not NaN_fec:
            session = Session()
            init = True
            # Traer todos los registros existentes de la base de datos
            fechasCompras = session.query(UltimaFechaCompra).all()
            fechas_dict = {(inv.REFERENCIA): inv for inv in fechasCompras}
            # Iterar sobre las filas del DataFrame
            for _, row in df.iterrows():
                clave = (str(int(row['REFERENCIA'])))
                if clave in fechas_dict:
                    # Si el registro existe, actualizar EXISTENCIA y ESTADO
                    fechasCompras = fechas_dict[clave]
                    fechasCompras.FECHA = str(row['FECHA'])
                else:
                    # Si no existe, crear un nuevo registro (nueva referencia)
                    nuevaFecha = UltimaFechaCompra(
                        REFERENCIA=str(int(row['REFERENCIA'])),
                        FECHA=str(row['FECHA'])
                    )
                    session.add(nuevaFecha)
            
            # Guardar los cambios en la base de datos
            session.commit()
                
    except Exception as e:
        print("Error en actualizar fecha ultima compra: ", e)
        session.rollback()  # Revertir en caso de error
    finally:
        if init:
            session.close()


# Caché para almacenar los resultados de las IPs resueltas
@lru_cache(maxsize=100)
def getName(ip):
    try:
        return socket.gethostbyaddr(ip)[0]  # Obtiene el nombre del equipo
    except socket.herror:
        return "Nombre desconocido"

def msg(client, mensaje):
    hora_actual = datetime.now()
    logFile = os.path.join(r'\\10.0.0.96\Compartida\logs', "peticiones.log")
    msg = f'[{getName(client)} ({client})][{hora_actual}]: {mensaje}'
    with open(logFile, "a") as log:
        log.write(f"{msg}\n")
    print(msg)

# METODO PARA BUSCAR UN INVENTARIO PARA UNA REFERENCIA
@controllerInventario.route('/inventario')
def inventarios():
    try:
        ref = request.args.get('r')
        resultados = []
        msg(request.remote_addr, f"Consulta inventario de ({ref})")
        if ref!="":
            session = Session()
            resultados = session.query(
                Referencia.REFERENCIA,
                Referencia.PRECIOXMAYOR,
                Referencia.PRECIOXUNIDAD,
                Referencia.DESCUENTO,
                Inventario.BODEGA,
                Inventario.LINEA,
                Inventario.EXISTENCIA,
                Inventario.COSTO
            ).join(Inventario, Referencia.REFERENCIA == Inventario.REFERENCIA) \
            .filter(Referencia.REFERENCIA == ref ) \
            .order_by(Inventario.BODEGA.asc()) \
            .all()
            json = [
            {
                'referencia': resultado.REFERENCIA,
                'precioxmayor': resultado.PRECIOXMAYOR,
                'precioxunidad': resultado.PRECIOXUNIDAD,
                'descuento': resultado.DESCUENTO,
                'bodega': resultado.BODEGA,
                'linea': resultado.LINEA,
                'existencia': resultado.EXISTENCIA,
                'costo': resultado.COSTO
            }
            for resultado in resultados
        ]

        return jsonify(json)
    except Exception as e:
        print(e)
    finally:
        if ref!="":
            session.close()
            
# CONTROLADORES PARA LA VISTA DE INVENTARIOS POR PROOVEDOR

@controllerInventario.route('/inventarios', methods=['GET'])
def vistaInventarios():
    msg(request.remote_addr, f"Entró a inventarios.")
    return render_template('inventarios.html')

@controllerInventario.route('/inv-consultas', methods=['GET'])
def listaInventarios():
    try:
        where = ""
        pro = request.args.get('p')
        sub = request.args.get('s')
        lin = request.args.get('l')
        ref = request.args.get('r')
        page = int(request.args.get('pg'))
        page_size = int(request.args.get('sz'))
        msg(request.remote_addr, f"Consulta inventarios [pro:{pro}][lin:{lin}][sub:{sub}][ref:{ref}][pag:{page}][tam:{page_size}]")
        offset = (page - 1) * page_size
        if pro != "" or sub != "" or lin != "" or ref != "":
            where = ""
            i = 0
            if pro != "":
                where += f"and i.PROVEEDOR = '{pro}'"
                i += 1
            if sub != "":
                where += f"and i.SUBGRUPO = '{sub}'"
                i += 1
            if lin != "":
                where += f"and i.LINEA = '{lin}'"
                i += 1
            if ref != "":
                where += f"and (lp.REFERENCIA LIKE '%{ref}%' or lp.DESCRIPCION LIKE '%{ref}%')"
                i += 1
                
        session = Session()
        total_rows = session.execute(text(f'''
                                                SELECT COUNT(DISTINCT lp.REFERENCIA)
                                                FROM ListaPreciosLectores as lp 
                                                INNER JOIN Inventarios as i
                                                    ON lp.REFERENCIA = i.REFERENCIA
                                                    {where};
                                            ''')).scalar()
        if offset >= total_rows:
            offset = total_rows
            
        inventarios = session.execute(text(f'''
                                          SELECT
                                                lp.REFERENCIA,
                                                lp.DESCRIPCION,
                                                SUM(CASE WHEN i.BODEGA = '1' THEN i.EXISTENCIA ELSE 0 END) AS 'BD1',
                                                SUM(CASE WHEN i.BODEGA = '2' THEN i.EXISTENCIA ELSE 0 END) AS 'BD2',
                                                SUM(CASE WHEN i.BODEGA = '20' THEN i.EXISTENCIA ELSE 0 END) AS 'BD20',
                                                lp.UNIDAD,
                                                lp.DESCUENTO,
                                                (SELECT CASE WHEN inv_b1.COSTO = 0 THEN ISNULL((SELECT inv_b2.COSTO FROM Inventarios AS inv_b2 WHERE inv_b2.REFERENCIA = lp.REFERENCIA AND inv_b2.BODEGA = '2'), 0) ELSE inv_b1.COSTO END FROM Inventarios AS inv_b1 WHERE inv_b1.REFERENCIA = lp.REFERENCIA AND inv_b1.BODEGA = '1') AS 'COSTO',
                                                lp.PRECIOXMAYOR,
                                                lp.PRECIOXUNIDAD,
                                                (SELECT FECHA FROM UltimaFechaCompra as fc WHERE fc.REFERENCIA = lp.REFERENCIA) as 'FECHACOMPRA'
                                            FROM ListaPreciosLectores as lp INNER JOIN Inventarios as i
                                                ON lp.REFERENCIA = i.REFERENCIA
                                                WHERE i.BODEGA IN ('1', '2', '20') {where}
                                            GROUP BY lp.REFERENCIA, lp.DESCRIPCION, lp.UNIDAD, lp.DESCUENTO, lp.PRECIOXMAYOR, lp.PRECIOXUNIDAD
                                            ORDER BY lp.DESCRIPCION ASC
                                            OFFSET {offset} ROWS FETCH NEXT {page_size} ROWS ONLY;
                                          ''')).fetchall()
        # lp.DESCRIPCION ASC
        session.close()
        json = [
            {
                'referencia': inventario[0],
                'descripcion': inventario[1],
                'bd1': inventario[2],
                'bd2': inventario[3],
                'bd20': inventario[4],
                'ue': inventario[5],
                'descuento': inventario[6],
                'costo': inventario[7],
                'precioxmayor': inventario[8],
                'precioxunidad': inventario[9],
                'fechacompra': inventario[10],
                
            }
            for inventario in inventarios
        ]
        
        resp = {
            "lista": json,
            "cantidad": total_rows
        }
        return jsonify(resp)
    except Exception as e:
        print(e)

#actualizarInv()

lineas = {10: "ALAMBRES", 7: "ALUMINIO Y FUNDIDO", 15: "BICICLETAS", 24: "DECORACION HOGAR", 1: "ELECTRODOMESTICOS MAYORES", 2: "ELECTRODOMESTICOS MENORES", 4: "ESCOLAR", 17: "HIERRO ALEADO Y FUNDIDO", 18: "HAMACAS Y TOALLAS", 5: "IMPORTADOS VARIOS", 6: "JUGUETERIA", 21: "JUGUETERIA IMPORTADA", 9: "LOCERIA", 14: "NACIONAL VARIOS", 22: "PIÑATERIA", 3: "PLASTICOS", 8: "CRISTALERIA", 19: "TELAS"}


@controllerInventario.route('/descargar_excel')
def descargar_excel():
    try:
        where = ""
        pro = request.args.get('p')
        sub = request.args.get('s')
        lin = request.args.get('l')
        ref = request.args.get('r')
        tp = request.args.get('tp')

        # Crear un archivo Excel en memoria usando pandas
        if pro!="" or sub!="" or lin!="" or ref!="":
            where = ""
            i = 0
            if pro != "":
                where += f"and i.PROVEEDOR = '{pro}'"
                i += 1
            if sub != "":
                where += f"and i.SUBGRUPO = '{sub}'"
                i += 1
            if lin != "":
                where += f"and i.LINEA = '{lin}'"
                i += 1
            if ref != "":
                where += f"and (lp.REFERENCIA LIKE '%{ref}%' or lp.DESCRIPCION LIKE '%{ref}%')"
                i += 1
        
        sql = f'''
            SELECT
                lp.REFERENCIA,
                lp.DESCRIPCION,
                (SELECT NOMBRE FROM Proveedor AS pr WHERE pr.PROVEEDOR = i.PROVEEDOR) AS 'PROVEEDOR',
                lp.UNIDAD AS "U/E",
                i.LINEA,
                SUM(CASE WHEN i.BODEGA = '1' THEN i.EXISTENCIA ELSE 0 END) AS 'BD1',
                SUM(CASE WHEN i.BODEGA = '2' THEN i.EXISTENCIA ELSE 0 END) AS 'BD2',
                -- SUM(CASE WHEN i.BODEGA = '3' THEN i.EXISTENCIA ELSE 0 END) AS 'BD3',
                -- SUM(CASE WHEN i.BODEGA = '7' THEN i.EXISTENCIA ELSE 0 END) AS 'BD7',
                SUM(CASE WHEN i.BODEGA = '20' THEN i.EXISTENCIA ELSE 0 END) AS 'BD20',
                lp.DESCUENTO,
                (SELECT CASE WHEN inv_b1.COSTO = 0 THEN ISNULL((SELECT inv_b2.COSTO FROM Inventarios AS inv_b2 WHERE inv_b2.REFERENCIA = lp.REFERENCIA AND inv_b2.BODEGA = '2'), 0) ELSE inv_b1.COSTO END FROM Inventarios AS inv_b1 WHERE inv_b1.REFERENCIA = lp.REFERENCIA AND inv_b1.BODEGA = '1') AS 'COSTO',
                lp.PRECIOXMAYOR,
                lp.PRECIOXUNIDAD,
                (SELECT FECHA FROM UltimaFechaCompra as fc WHERE fc.REFERENCIA = lp.REFERENCIA) as 'FECHACOMPRA'
            FROM ListaPreciosLectores as lp INNER JOIN Inventarios as i
                ON lp.REFERENCIA = i.REFERENCIA
                WHERE i.BODEGA IN ('1', '2', '20') {where}
            GROUP BY lp.REFERENCIA, lp.DESCRIPCION, lp.UNIDAD, i.LINEA, lp.DESCUENTO, lp.PRECIOXMAYOR, lp.PRECIOXUNIDAD, i.PROVEEDOR
            ORDER BY lp.DESCRIPCION ASC
        '''

        df = pd.read_sql(sql, engine)
        df['LINEA'] = df['LINEA'].map(lineas)
        df['Renta x Mayor'] = df.apply(lambda x: (1 - (x['COSTO'] / x['PRECIOXMAYOR'])) if x['PRECIOXMAYOR'] > 0 else None, axis=1)
        df['Renta x Unidad'] = df.apply(lambda x: (1 - (x['COSTO'] / x['PRECIOXUNIDAD'])) if x['PRECIOXUNIDAD'] > 0 else None, axis=1)
        df['FECHACOMPRA'] = pd.to_datetime(df['FECHACOMPRA'], errors='coerce')
        df['FECHACOMPRA'] = df['FECHACOMPRA'].dt.strftime('%b %d %Y').str.replace('.', '', regex=False).str.capitalize()
        
        columnas_ordenadas = [
            'PROVEEDOR', 'REFERENCIA', 'DESCRIPCION', 'LINEA',  'U/E', 'BD1', 'BD2', 'BD20', 'DESCUENTO', 
            'COSTO', 'PRECIOXMAYOR', 'PRECIOXUNIDAD', 'Renta x Mayor', 'Renta x Unidad', 'FECHACOMPRA'
        ]
        if tp == "@inv":
            columnas_ordenadas = ['PROVEEDOR', 'REFERENCIA', 'DESCRIPCION', 'LINEA',  'U/E', 'BD1', 'BD2', 'BD20']
        elif tp == "@all":
            columnas_ordenadas = [
            'PROVEEDOR', 'REFERENCIA', 'DESCRIPCION', 'LINEA',  'U/E', 'BD1', 'BD2', 'BD3', 'BD7', 'BD20', 'DESCUENTO', 
            'COSTO', 'PRECIOXMAYOR', 'PRECIOXUNIDAD', 'Renta x Mayor', 'Renta x Unidad', 'FECHACOMPRA'
        ]

        # Reordenar el DataFrame
        df = df[columnas_ordenadas]
        excel_io = io.BytesIO()

        # Convertir el DataFrame a un archivo Excel con pandas
        with pd.ExcelWriter(excel_io, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Datos')

            # Obtener el objeto del libro y la hoja
            workbook = writer.book
            sheet = workbook['Datos']

            # Ajustar el tamaño de las columnas (autoajuste básico)
            for col in sheet.columns:
                max_length = 0
                column = col[0].column_letter  # Get the column name
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 10)
                sheet.column_dimensions[column].width = adjusted_width

            # Aplicar color gris claro a los encabezados
            header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # Gris claro
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')

            if tp != "@inv":
                # Aplicar formato de contabilidad (moneda) a las columnas Costo, PrecioXMAYOR y PrecioXUNIDAD
                for row in sheet.iter_rows(min_row=2, min_col=10, max_col=12):  # Columnas Costo, PrecioXMAYOR, PrecioXUNIDAD
                    for cell in row:
                        cell.number_format = '_("$"* #,##0.00'  # Formato de contabilidad
                
                for row in sheet.iter_rows(min_row=2, min_col=13, max_col=14):  # Columnas Renta x Mayor, Renta x Unidad
                            for cell in row:
                                cell.number_format = '0.00%'  # Formato de porcentaje con 2 decimales
        # Mover el cursor al principio del archivo
        excel_io.seek(0)

        # Enviar el archivo Excel como respuesta
        return send_file(
            excel_io,
            as_attachment=True,
            download_name=f"INVENTARIO-{str(datetime.now()).replace('-','').replace('.','')}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        print(e)
    
@controllerInventario.route('/ultimacompra')
def generarUltimaCompra():
    try:
        df = pd.read_sql("SELECT * FROM UltimaFechaCompra", engine)
        df['REFERENCIA'] = pd.to_numeric(df['REFERENCIA'], errors='coerce').astype('Int64')
        df['FECHA'] = pd.to_datetime(df['FECHA'], errors='coerce')
        df['FECHA'] = df['FECHA'].dt.strftime('%b %d %Y').str.replace('.', '', regex=False).str.capitalize()
        df = df[['REFERENCIA', 'FECHA']]

        ruta_salida = r'I:\Iltda\Soporte Compartida\Compras\Base Compras\Compras\FECHA ULTIMA COMPRA.xlsx'

        df.to_excel(ruta_salida, index=False, sheet_name='Datos')

        wb = load_workbook(ruta_salida)
        ws = wb['Datos']
        for cell in ws['A'][1:]:  # omite el encabezado
            cell.number_format = numbers.FORMAT_NUMBER 
        max_row = ws.max_row
        rango = DefinedName('rango', attr_text=f'Datos!$A$1:$B${max_row}')
        wb.defined_names.add(rango)

        wb.save(ruta_salida)

        print("archivo FECHA ULTIMA COMPRA.xlsx generado correctamente.")
        return "archivo FECHA ULTIMA COMPRA.xlsx generado correctamente."
    except Exception as e:
        print("Error al generar FECHA ULTIMA COMPRA.xlsx: ", e)
        
def actualizaciones():
    try:
        actualizarInv()
        actualizarFechaCompra()
    except Exception as e:
        print("Error al actualizar: ", e)


scheduler = BackgroundScheduler()
def iniciarCargue():
    try:
        existing_job = scheduler.get_job('actualizaciones_id')
        if not existing_job:
            scheduler.add_job(
                actualizaciones,
                'interval',
                seconds=40,
                id='actualizaciones_id',
                max_instances=1,
                coalesce=True
            )

            scheduler.start()
    except Exception as e:
        print(e)


iniciarCargue()